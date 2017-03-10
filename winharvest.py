"""
Query for installed software on a Windows system
- should get the same list as what is displayed in Control Panel + more Windows items (updates, Office, etc.)
- checks for architecture to know what registry keys to search
- writes Excel file with results
"""

import os
import socket
from _winreg import *
import platform
import openpyxl
from openpyxl.styles import Font


# some constants here
SYSTEM = HKEY_LOCAL_MACHINE
USER = HKEY_CURRENT_USER
PATH = r"SOFTWARE\Microsoft\Windows\CurrentVersion"
KEYNAME = 'Uninstall'
SOFTWARE = r"SOFTWARE\Microsoft\Windows\CurrentVersion\Uninstall"

def key_exists(hive, keypath, name):
    '''
    Checks if key exists in the searched path
    :param hive: one of the main hives to open
    :param keypath: the path to look for the key
    :param name: the name of the key to search for
    :return: True if the key was found, False otherwise
    '''

    reg = OpenKey(hive, keypath)
    total = QueryInfoKey(reg)[0]
    for k in range(0, total + 1):
        try:
            kname = EnumKey(reg, k)
            if kname == name:
                return True
        except WindowsError:
            break
    return False

def key_exists64(hive, keypath, name):
    '''
    Checks if key exists in the searched path (for the 64 bit registry)
    :param hive: one of the main hives to open
    :param keypath: the path to look for the key
    :param name: the name of the key to search for
    :return: True if the key was found, False otherwise
    '''

    reg = OpenKey(hive, keypath, 0, KEY_READ)
    total = QueryInfoKey(reg)[0]
    for k in range(0, total + 1):
        try:
            kname = EnumKey(reg, k)
            if kname == name:
                return True
        except WindowsError:
            break
    return False

apps = {} # list of installed programs
def get_reg_apps(handle):
    '''
    Enumerate the installed programs
    :param handle: handle to registry
    :return:
    '''
    # connect to registry handle
    total = QueryInfoKey(handle)[0] # the 1st item of the 3 in tuple is the no. of subkeys for this key
    for key in range(0, total + 1): # loop over all keys for installed software
        try:
            keyname = EnumKey(handle, key) # get the name of the subkey
            subkey = OpenKey(handle, keyname) # open handle to subkey
            try:
                name = QueryValueEx(subkey, 'DisplayName')[0] # get the value of the DisplayName property
                try:
                    version = QueryValueEx(subkey, 'DisplayVersion')[0]
                except WindowsError, e:
                    version = ''
                    pass

                if name != WindowsError and name not in apps and 'Security Update for Microsoft' not in name and \
                        'Update for Microsoft' not in name:
                            apps[name] = version
            except WindowsError, e:
                # print e # DEBUG
                CloseKey(subkey)
                pass
        except WindowsError, e: # no more keys to read
            # print e # DEBUG
            # clean up and close handles
            pass
        #CloseKey(subkey)
    CloseKey(handle)


arch = platform.machine() # can be 'AMD64' or 'x86'

if key_exists(SYSTEM, PATH, KEYNAME):
    regkey = OpenKey(SYSTEM, SOFTWARE, 0, KEY_READ) # open the key identified in 2nd param
    get_reg_apps(regkey)
if key_exists(USER, PATH, KEYNAME):
    curr_user = OpenKey(USER, SOFTWARE, 0, KEY_READ)
    get_reg_apps(curr_user)
if arch == 'AMD64':
    if key_exists(SYSTEM, PATH, KEYNAME):
        regkey64 = OpenKey(SYSTEM, SOFTWARE, 0, KEY_READ | KEY_WOW64_64KEY) # per MSDN, it must be ORed with the access flag
        get_reg_apps(regkey64)
    if key_exists(USER, PATH, KEYNAME):
        curr_user64 = OpenKey(USER, SOFTWARE, 0, KEY_READ | KEY_WOW64_64KEY)
        get_reg_apps(curr_user64)


print 'Number of installed programs: ' +  str(len(apps))


def get_ip():
    """
    Get the IP of a working interface
    :return:
    """
    ip = socket.gethostbyname(socket.gethostname())
    if ip.startswith('169'):
        try:
            s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
            s.connect(('8.8.8.8', 53))  # UDP connection to Google DNS
            working_ip = s.getsockname()[0] # get only the IP of the socket
            return working_ip
        except socket.error, e:
            print e
            return ip
    else:
        return ip

host =  os.environ['COMPUTERNAME']
username = os.getenv('username')
ip = get_ip()
filename = host + '.txt'

FILE = host + '.xlsx'

def write_xls():
    '''
    Create an Excel file with the desired values
    :return:
    '''
    wbook = openpyxl.Workbook()
    sheet = wbook.get_sheet_by_name('Sheet')
    sheet['A1'] = 'COMPUTER NAME'
    sheet.column_dimensions['A'].width = 25
    sheet['B1'] = 'USER'
    sheet.column_dimensions['B'].width = 25
    sheet['C1'] = 'IP'
    sheet.column_dimensions['C'].width = 20
    sheet['D1'] = 'INSTALLED SOFTWARE'
    sheet.column_dimensions['D'].width = 80
    sheet['E1'] = 'VERSION'
    sheet.column_dimensions['E'].width = 20
    sheet['F1'] = 'TOTAL'
    wbook.save(FILE)
    return

write_xls()

def boldfont(spreadsheet):
    '''
    Apply bold font
    :param spreadsheet: Excel file to open and make changes to
    :return:
    '''
    wbook = openpyxl.load_workbook(spreadsheet)
    sheet = wbook.get_sheet_by_name('Sheet')
    bold = Font(name='Times New Roman', bold=True) # bold font
    for row in sheet.iter_rows():
        for cell in row:
            cell.font = bold # apply font
    wbook.save(spreadsheet)
    return

boldfont(FILE)

def write_to_excel(spreadsheet):
    '''
    Write values
    :param spreadsheet: Excel file to open and make changes to
    :return:
    '''
    wbook = openpyxl.load_workbook(spreadsheet)
    sheet = wbook.get_sheet_by_name('Sheet')
    sheet['A2'] = host
    sheet['B2'] = username
    sheet['C2'] = ip
    start_count = 2
    for app, version in apps.iteritems():
        sheet['D' + str(start_count)] = app.encode('utf-8')
        sheet['E' + str(start_count)] = version.encode('utf-8')
        start_count += 1
    sheet['F2'] = len(apps)
    wbook.save(spreadsheet)
    return

write_to_excel(FILE)




